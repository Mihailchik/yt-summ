"""
Модуль для сохранения результатов в Excel файл
"""
import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    load_workbook = None
    get_column_letter = None


# Константы заголовков колонок (строго не менять порядок и названия)
REQUIRED_COLUMNS = [
    "Номер",
    "Ссылка", 
    "Субтитры",
    "Чистый текст",
    "Фулл саммари",
    "Мидл саммари",
    "Шорт саммари",
    "Материалы",
    "Дата добавления"
]


class ExcelHandle:
    """Объект для работы с Excel файлом"""
    def __init__(self, workbook, worksheet, file_path, column_mapping, config):
        self.workbook = workbook
        self.worksheet = worksheet
        self.file_path = file_path
        self.column_mapping = column_mapping  # {column_name: column_index}
        self.config = config


def init_excel(config, log):
    """
    Гарантирует наличие файла и листа.
    Если файла нет — создать.
    Если листа нет — создать.
    Если каких-то колонок нет — добавить их В КОНЕЦ. Ничего не удалять и не переименовывать.
    Вернуть handle с данными (wb, ws, путь, имена/индексы колонок).
    Залогировать: excel_init file=... sheet=... created={true|false} added_cols=[...]
    """
    
    if load_workbook is None or Workbook is None:
        log("ERROR", "store_excel", "openpyxl не установлен. Запустите: pip install openpyxl")
        raise ValueError("openpyxl_not_installed")
    
    excel_config = config.get('excel', {})
    file_path = excel_config.get('file_path', 'config/yt_summary.xlsx')
    sheet_name = excel_config.get('sheet_name', 'YT_SUM')
    
    # Обеспечиваем абсолютный путь относительно директории проекта
    if not os.path.isabs(file_path):
        # Получаем путь к директории проекта (родительская директория от src)
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        file_path = os.path.join(project_dir, file_path)
    
    # Создаем директорию если не существует
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    file_created = False
    sheet_created = False
    added_cols = []
    
    try:
        # Пробуем загрузить существующий файл
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            log("DEBUG", "store_excel", "Загружен существующий Excel файл", file=file_path)
        else:
            # Создаем новый файл
            workbook = Workbook()
            file_created = True
            log("DEBUG", "store_excel", "Создан новый Excel файл", file=file_path)
        
        # Проверяем наличие листа
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            log("DEBUG", "store_excel", "Использован существующий лист", sheet=sheet_name)
        else:
            # Создаем новый лист
            if file_created and 'Sheet' in workbook.sheetnames:
                # Переименовываем дефолтный лист
                workbook['Sheet'].title = sheet_name
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)
            sheet_created = True
            log("DEBUG", "store_excel", "Создан новый лист", sheet=sheet_name)
        
        # Анализируем существующие заголовки
        existing_headers = []
        if worksheet.max_row >= 1:
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    existing_headers.append(str(cell_value))
                else:
                    break
        
        # Создаем mapping существующих колонок
        column_mapping = {}
        for i, header in enumerate(existing_headers):
            column_mapping[header] = i + 1
        
        # Добавляем недостающие колонки
        next_col = len(existing_headers) + 1
        for required_col in REQUIRED_COLUMNS:
            if required_col not in column_mapping:
                worksheet.cell(row=1, column=next_col, value=required_col)
                column_mapping[required_col] = next_col
                added_cols.append(required_col)
                next_col += 1
        
        # Сохраняем файл если были изменения
        if file_created or sheet_created or added_cols:
            workbook.save(file_path)
        
        log("INFO", "store_excel", "Excel инициализирован", 
            file=file_path, sheet=sheet_name, created=file_created or sheet_created, 
            added_cols=added_cols)
        
        return ExcelHandle(workbook, worksheet, file_path, column_mapping, excel_config)
        
    except PermissionError as e:
        log("ERROR", "store_excel", "Excel файл заблокирован", 
            action="init_excel", reason=str(e), file=file_path)
        raise ValueError("excel_file_locked")
    except Exception as e:
        log("ERROR", "store_excel", "Ошибка инициализации Excel", 
            action="init_excel", reason=str(e))
        raise ValueError(f"excel_init_error: {e}")


def allocate_run_id(handle, log) -> int:
    """
    Найти максимальный 'Номер' в листе (игнорируя шапку и пустые строки).
    Вернуть max+1, если пусто — 1.
    Зарезервировать строку под этот номер (создать запись с 'Номер' и 'Дата добавления'=сейчас в ISO-8601).
    Залогировать: excel_allocate_run_id id=<N>
    """
    
    try:
        worksheet = handle.worksheet
        column_mapping = handle.column_mapping
        
        if "Номер" not in column_mapping:
            log("ERROR", "store_excel", "Колонка 'Номер' не найдена")
            raise ValueError("column_not_found")
        
        number_col = column_mapping["Номер"]
        max_number = 0
        
        # Ищем максимальный номер (начинаем с 2й строки, т.к. 1я - заголовки)
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=number_col).value
            if cell_value is not None:
                try:
                    number = int(cell_value)
                    max_number = max(max_number, number)
                except (ValueError, TypeError):
                    continue
        
        # Новый ID
        run_id = max_number + 1
        
        # Находим первую свободную строку
        new_row = worksheet.max_row + 1
        if worksheet.max_row == 1:  # Только заголовки
            new_row = 2
        
        # Записываем номер
        worksheet.cell(row=new_row, column=number_col, value=run_id)
        
        # Записываем дату добавления (без миллисекунд)
        if "Дата добавления" in column_mapping:
            date_col = column_mapping["Дата добавления"]
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=new_row, column=date_col, value=current_time)
        
        # Сохраняем файл
        handle.workbook.save(handle.file_path)
        
        log("INFO", "store_excel", "Выделен ID для записи", 
            excel_allocate_run_id=True, id=run_id)
        
        return run_id
        
    except Exception as e:
        log("ERROR", "store_excel", "Ошибка выделения ID", 
            action="allocate_run_id", reason=str(e))
        raise ValueError(f"allocate_run_id_error: {e}")


def write_step(handle, run_id: int, updates: dict, log) -> None:
    """
    Обновить только переданные поля в строке с данным 'Номер'.
    Поддерживаемые ключи: 'Ссылка', 'Субтитры', 'Чистый текст', 'Дата добавления'.
    Перед записью в 'Субтитры' и 'Чистый текст' применить тримминг: len>max_cell_chars → text[:max_cell_chars].
    Сохранить файл.
    Лог: excel_write_step id=<N> fields=[...] (если 'Субтитры'/'Чистый текст' триммированы — truncated=true orig_len=... saved_len=32767)
    """
    
    try:
        worksheet = handle.worksheet
        column_mapping = handle.column_mapping
        max_cell_chars = handle.config.get('max_cell_chars', 32767)
        
        if "Номер" not in column_mapping:
            log("ERROR", "store_excel", "Колонка 'Номер' не найдена")
            raise ValueError("column_not_found")
        
        number_col = column_mapping["Номер"]
        target_row = None
        
        # Находим строку с нужным номером
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=number_col).value
            if cell_value is not None:
                try:
                    if int(cell_value) == run_id:
                        target_row = row
                        break
                except (ValueError, TypeError):
                    continue
        
        if target_row is None:
            log("ERROR", "store_excel", "Строка с указанным ID не найдена", 
                run_id=run_id)
            raise ValueError(f"run_id_not_found: {run_id}")
        
        # Обновляем поля
        updated_fields = []
        truncated = False
        orig_len = None
        saved_len = None
        
        for field_name, field_value in updates.items():
            if field_name not in column_mapping:
                log("WARN", "store_excel", f"Колонка '{field_name}' не найдена, пропускаем")
                continue
            
            col_index = column_mapping[field_name]
            value_to_write = field_value
            
            # Применяем тримминг для текстовых полей
            if field_name in ["Субтитры", "Чистый текст"] and isinstance(field_value, str):
                orig_len = len(field_value)
                if orig_len > max_cell_chars:
                    value_to_write = field_value[:max_cell_chars]
                    truncated = True
                    saved_len = len(value_to_write)
            
            worksheet.cell(row=target_row, column=col_index, value=value_to_write)
            updated_fields.append(field_name)
        
        # Сохраняем файл
        handle.workbook.save(handle.file_path)
        
        # Логируем результат
        log_params = {
            "excel_write_step": True,
            "id": run_id,
            "fields": updated_fields
        }
        
        if truncated:
            log_params.update({
                "truncated": True,
                "orig_len": orig_len,
                "saved_len": saved_len
            })
        
        log("INFO", "store_excel", "Обновлены поля записи", **log_params)
        
    except PermissionError as e:
        log("ERROR", "store_excel", "Excel файл заблокирован при записи", 
            action="write_step", reason=str(e), run_id=run_id)
        raise ValueError("excel_file_locked")
    except Exception as e:
        log("ERROR", "store_excel", "Ошибка записи полей", 
            action="write_step", reason=str(e), run_id=run_id)
        raise ValueError(f"write_step_error: {e}")


def close(handle):
    """
    Закрыть ресурсы, если требуется выбранной библиотекой.
    """
    try:
        # openpyxl не требует явного закрытия, но сохраним файл на всякий случай
        if handle and handle.workbook:
            handle.workbook.save(handle.file_path)
        # Обнуляем ссылки
        if handle:
            handle.workbook = None
            handle.worksheet = None
    except Exception:
        # Игнорируем ошибки при закрытии
        pass


# =================== ТЕСТОВЫЕ ФУНКЦИИ ===================
# Эти функции будут удалены после тестирования

def init_test_sheet(handle, log):
    """
    Создает тестовый лист 'AI_TESTS' для диагностики AI вызовов.
    Колонки: ID попытки, Успешная, Модель AI, Время Supadata (мс), Время AI (мс), Дата
    """
    try:
        sheet_name = "AI_TESTS"
        workbook = handle.workbook
        
        # Создаем лист если не существует
        if sheet_name not in workbook.sheetnames:
            test_sheet = workbook.create_sheet(sheet_name)
            
            # Заголовки колонок
            headers = [
                "ID попытки",
                "Успешная", 
                "Модель AI",
                "Время Supadata (мс)",
                "Время AI (мс)",
                "Символов исходно",
                "Символов после AI",
                "Сжатие (%)",
                "Дата"
            ]
            
            for col, header in enumerate(headers, 1):
                test_sheet.cell(row=1, column=col, value=header)
                
            workbook.save(handle.file_path)
            log("INFO", "store_excel", "Создан тестовый лист AI_TESTS")
        else:
            test_sheet = workbook[sheet_name]
            log("DEBUG", "store_excel", "Используется существующий лист AI_TESTS")
            
        return test_sheet
        
    except Exception as e:
        log("ERROR", "store_excel", "Ошибка создания тестового листа", error=str(e))
        return None


def write_test_record(handle, test_data: dict, log):
    """
    Записывает данные теста в лист AI_TESTS.
    test_data = {
        'run_id': int,
        'success': bool,
        'model': str,
        'supadata_time_ms': int,
        'ai_time_ms': int,
        'chars_original': int,     # новое поле
        'chars_cleaned': int,      # новое поле
        'compression_percent': float  # новое поле
    }
    """
    try:
        test_sheet = init_test_sheet(handle, log)
        if test_sheet is None:
            return
            
        # Находим следующую свободную строку
        next_row = test_sheet.max_row + 1
        
        # Вычисляем процент сжатия
        chars_original = test_data.get('chars_original', 0)
        chars_cleaned = test_data.get('chars_cleaned', 0)
        
        if chars_original > 0:
            compression_percent = round(((chars_original - chars_cleaned) / chars_original) * 100, 1)
        else:
            compression_percent = 0.0
        
        # Записываем данные
        test_sheet.cell(row=next_row, column=1, value=test_data.get('run_id', 0))
        test_sheet.cell(row=next_row, column=2, value="ДА" if test_data.get('success', False) else "НЕТ")
        test_sheet.cell(row=next_row, column=3, value=test_data.get('model', 'unknown'))
        test_sheet.cell(row=next_row, column=4, value=test_data.get('supadata_time_ms', 0))
        test_sheet.cell(row=next_row, column=5, value=test_data.get('ai_time_ms', 0))
        test_sheet.cell(row=next_row, column=6, value=chars_original)  # новая колонка
        test_sheet.cell(row=next_row, column=7, value=chars_cleaned)   # новая колонка
        test_sheet.cell(row=next_row, column=8, value=compression_percent)  # новая колонка
        test_sheet.cell(row=next_row, column=9, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Сохраняем файл
        handle.workbook.save(handle.file_path)
        
        log("INFO", "store_excel", "Записан тест в AI_TESTS", 
            run_id=test_data.get('run_id'), 
            success=test_data.get('success'),
            model=test_data.get('model', 'unknown'),
            chars_original=chars_original,
            chars_cleaned=chars_cleaned,
            compression=f"{compression_percent}%")
            
    except Exception as e:
        log("ERROR", "store_excel", "Ошибка записи тестовых данных", error=str(e))